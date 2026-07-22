import csv
import io
from datetime import date, datetime
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Human-readable headers + column order (2026-07-15) -- these used to be the
# raw column names (customer_id, responsible_user_id, ...), so an ordinary
# admin opening the file in Excel saw a wall of UUIDs/integers with nothing
# to explain them, easily mistaken for corrupted/"encrypted" data. Now: a
# readable header label per column, names instead of ids (see
# reports/sql/queries.sql's export_* queries, which now join for these), and
# the row-id UUID moved last since it's only useful for support lookups, not
# reading the report.
CUSTOMERS_COLUMNS = [
    ("full_name", "F.I.Sh."),
    ("phone", "Telefon"),
    ("stage", "Bosqich"),
    ("responsible_user_name", "Mas'ul xodim"),
    ("created_at", "Yaratilgan sana"),
    ("updated_at", "Yangilangan sana"),
    ("id", "ID"),
]
SALES_COLUMNS = [
    ("customer_name", "Mijoz"),
    ("customer_phone", "Mijoz telefoni"),
    ("category_name", "Kategoriya"),
    ("responsible_user_name", "Mas'ul xodim"),
    ("price_amount", "Summa"),
    ("deadline", "Muddat"),
    ("status", "Holat"),
    ("created_at", "Yaratilgan sana"),
    ("id", "ID"),
]
FINANCE_COLUMNS = [
    ("customer_name", "Mijoz"),
    ("entry_type", "Yozuv turi"),
    ("amount", "Summa"),
    ("description", "Izoh"),
    ("created_at", "Sana"),
    ("id", "ID"),
]
CALLS_COLUMNS = [
    ("responsible_user_name", "Mas'ul xodim"),
    ("direction", "Yo'nalish"),
    ("from_number", "Kimdan"),
    ("to_number", "Kimga"),
    ("duration_seconds", "Davomiyligi (soniya)"),
    ("status", "Holat"),
    ("started_at", "Boshlangan vaqti"),
    ("ended_at", "Tugagan vaqti"),
    ("provider", "Provayder"),
    ("id", "ID"),
]

_MONEY_COLUMNS = {"price_amount", "amount"}
# Bug found 2026-07-15: a long all-digit string like "+998902598836" is
# exactly what Excel's CSV import heuristics treat as a number -- it then
# renders in scientific notation (9,98903E+11) and silently loses precision.
# These columns are never meant to be arithmetic, so they're forced to text.
_TEXT_FORCED_COLUMNS = {"phone", "customer_phone", "from_number", "to_number"}
_MAX_COLUMN_WIDTH = 40


def _format_money(amount: int, currency: str | None) -> str:
    # Same USD-is-cents convention as frontend/src/lib/format/money.ts --
    # never show a raw integer next to a currency code without converting.
    value = amount / 100 if currency == "USD" else amount
    formatted = f"{value:,.2f}" if currency == "USD" else f"{value:,.0f}"
    return f"{formatted} {currency}" if currency else str(amount)


def _format_datetime(value: datetime | date) -> str:
    # Bug found 2026-07-15: str(datetime) included microseconds and a
    # "+00:00" UTC offset suffix ("2026-07-14 04:42:23.242519+00:00") --
    # unreadable noise for a report a person is meant to scan. Trimmed to
    # what's actually useful: "2026-07-14 04:42:23".
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value.strftime("%Y-%m-%d")


def _cell_value(row: dict, column: str) -> object:
    value = row.get(column)
    if column in _MONEY_COLUMNS and value is not None:
        return _format_money(value, row.get("currency"))
    if isinstance(value, (datetime, date)):
        return _format_datetime(value)
    # UUID isn't natively writable by csv/openpyxl -- stringify it, same
    # reasoning as core/database's json.dumps(default=str) for JSONB.
    if isinstance(value, UUID):
        return str(value)
    return value


def rows_to_csv(rows: list[dict], columns: list[tuple[str, str]]) -> str:
    # Bug found 2026-07-15: plain comma-delimited CSV opened as one giant
    # unsplit column in Excel -- Excel's CSV import delimiter follows the
    # OS/Office regional "list separator" setting, which is ";" (not ",")
    # under Russian/most CIS locales (comma is that locale's decimal point
    # instead). ";" is the delimiter Excel actually expects there, so use it
    # rather than assuming a US locale. The leading ﻿ (UTF-8 BOM) makes
    # Excel auto-detect UTF-8 instead of guessing a legacy codepage --
    # otherwise Uzbek/Cyrillic names can render as mojibake.
    buffer = io.StringIO()
    buffer.write("﻿")
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow([label for _, label in columns])
    for row in rows:
        values = []
        for column, _ in columns:
            value = _cell_value(row, column)
            # ="..." is the standard CSV trick to force Excel to treat a
            # value as literal text instead of guessing a number -- a plain
            # quoted string doesn't survive Excel's own type-sniffing on CSV
            # import (unlike a real .xlsx cell, which carries an explicit
            # type and doesn't need this).
            if column in _TEXT_FORCED_COLUMNS and value:
                value = f'="{value}"'
            values.append(value)
        writer.writerow(values)
    return buffer.getvalue()


def rows_to_xlsx(rows: list[dict], columns: list[tuple[str, str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([label for _, label in columns])
    for row in rows:
        sheet.append([_cell_value(row, column) for column, _ in columns])

    # Bold header + frozen top row (2026-07-15) -- so a long export stays
    # readable while scrolling instead of losing track of which column is which.
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    for idx, (column, label) in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        if column in _TEXT_FORCED_COLUMNS:
            # Belt-and-suspenders: a real .xlsx cell already carries its own
            # type (openpyxl writes these as strings, so Excel won't
            # auto-number them the way it does with CSV) -- number_format='@'
            # additionally pins the column to Text so pasting/re-entry can't
            # reintroduce the same scientific-notation bug.
            for cell in sheet[letter]:
                cell.number_format = "@"
        widest = max((len(str(_cell_value(row, column) or "")) for row in rows), default=0)
        # +4 padding (not +2) -- explicitly asked for more breathing room
        # between columns than a bare best-fit width gives.
        sheet.column_dimensions[letter].width = min(max(len(label), widest) + 4, _MAX_COLUMN_WIDTH)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
